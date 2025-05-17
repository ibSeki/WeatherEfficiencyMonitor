import psycopg2
from psycopg2.extras import RealDictCursor
from datetime import datetime
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os
import pandas as pd
from io import BytesIO

DB_NAME = os.getenv('DB_NAME')
DB_USER = os.getenv('DB_USER')
DB_PASSWORD = os.getenv('DB_PASSWORD')
DB_HOST = os.getenv('DB_HOST')
DB_PORT = os.getenv('DB_PORT')

def get_connection():
    print(f"Conectando ao banco: {DB_NAME}@{DB_HOST}:{DB_PORT} como {DB_USER}")
    return psycopg2.connect(
        dbname=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        host=DB_HOST,
        port=DB_PORT
    )

def insert_weather_data(city, temperature, efficiency):
    sql = """
        INSERT INTO current_weather (city, temperature, efficiency, datetime)
        VALUES (%s, %s, %s, %s)
    """
    conn = get_connection()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(sql, (city, temperature, efficiency, datetime.utcnow()))
    finally:
        conn.close()

def get_weather_history():
    sql = """
        SELECT id, city, temperature, efficiency, datetime FROM current_weather ORDER BY datetime DESC LIMIT 100
    """
    conn = get_connection()
    try:
        with conn:
            with conn.cursor(cursor_factory=RealDictCursor) as cur:
                cur.execute(sql)
                results = cur.fetchall()
                return results
    finally:
        conn.close()

def clear_weather_history():
    sql = "DELETE FROM current_weather"
    conn = get_connection()
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute(sql)
    finally:
        conn.close()

def generate_excel():
    history = get_weather_history()
    if not history:
        return None

    df = pd.DataFrame(history)
    df = df[['city', 'temperature', 'efficiency', 'datetime']]

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Histórico')
        workbook  = writer.book
        worksheet = writer.sheets['Histórico']


        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, col_letter in enumerate(worksheet.iter_cols(min_row=1, max_row=1), 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment


        for i, column_cells in enumerate(worksheet.columns, 1):
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(i)].width = adjusted_width


        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=2, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal='right')

    output.seek(0)
    return output